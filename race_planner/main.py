"""
Main entry point for trail race planner.

Usage:
    python -m race_planner.main <race_config.yaml> [options]

Options:
    --athlete NAME             Athlete name (default: yet_another_sato)
    --mode MODE                Planning mode (default: athlete_pb):
                                 athlete_pb   — derive pace from athlete reference PB
                                 target_time  — plan to a given total finish time
                                 target_itra  — plan to a given target ITRA score
                                 grade_adjusted_pace — plan from a GAP-weighted pace
    --target-time HH:MM:SS     Required for --mode target_time
    --target-itra-score N      Required for --mode target_itra
    --target-grade-adjusted-pace MM:SS
                               Required for --mode grade_adjusted_pace
    --fatigue-mode {none|athlete|race}
                               Fatigue model source (default: none)
    --fatigue-total-decay-pct PCT
                               Override fatigue with linear decay (0–100); takes precedence
    --nutrition {yes|no}      Include nutrition column in main HTML report (default: no)

Notes:
    - For target_time, the provided time is the desired TOTAL finish time
      (running + all aid-station stops).
    - For target_itra, the race YAML must contain at least one entry in
      race.itra_reference_points.
    - The pacing plan is written as a new sheet in the existing
      segment-analysis Excel file defined by race.output_file.
"""

import argparse
import sys
from pathlib import Path

import yaml
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from race_planner.course import analyze_course
from race_planner.models.itra_predictor import ItraScorePredictor
from race_planner.models.nutrition import build_race_nutrition_plan, load_food_catalog
from race_planner.models.tools import (
    canonical_point_name,
    extract_volume_ml,
    format_decimal_quantity,
    hms_to_seconds,
    hours_to_hms,
    pace_to_seconds_per_km,
    seconds_to_hms,
)
from race_planner.planner import PaceCalculator
from race_planner.visualization.race_plan_table import generate_race_plan_table_report


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _total_stop_time_s(aid_stations: list) -> float:
    return sum(float(aid.get("stop_time_s", 0)) for aid in aid_stations)


def _resolve_fatigue_total_decay_pct(
    fatigue_mode: str,
    fatigue_total_decay_pct_cli: float | None,
    race_config: dict,
    athlete_config: dict,
) -> float:
    """
    Resolve fatigue total decay percentage with precedence:
    CLI override > race config > athlete config > 0 (default)

    Args:
        fatigue_mode: "none", "athlete", or "race"
        fatigue_total_decay_pct_cli: Optional CLI override
        race_config: Loaded race YAML
        athlete_config: Loaded athlete YAML

    Returns:
        Decay percentage (0-100), or 0 if mode is "none"
    """
    # CLI override takes absolute precedence
    if fatigue_total_decay_pct_cli is not None:
        return float(fatigue_total_decay_pct_cli)

    # If mode is "none", always 0
    if fatigue_mode == "none":
        return 0.0

    # Mode "race": try race config planning section
    if fatigue_mode == "race":
        planning = race_config.get("race", {}).get("planning", {})
        decay = planning.get("fatigue_total_decay_pct")
        if decay is not None:
            return float(decay)
        logger.warning(
            "--fatigue-mode race but no race.planning.fatigue_total_decay_pct found; defaulting to 0"
        )
        return 0.0

    # Mode "athlete": try athlete config (once physiological params are designed)
    if fatigue_mode == "athlete":
        # TODO: implement when process-based model is designed
        logger.warning("--fatigue-mode athlete not yet implemented; defaulting to 0")
        return 0.0

    return 0.0


def _build_itra_predictor(race_config: dict) -> ItraScorePredictor | None:
    """Return an ItraScorePredictor from the first itra_reference_points entry, or None."""
    ref_points = race_config.get("race", {}).get("itra_reference_points", [])
    if not ref_points:
        logger.warning(
            "No 'itra_reference_points' in race config — " "ITRA score prediction unavailable."
        )
        return None
    ref = ref_points[0]
    try:
        return ItraScorePredictor(
            reference_time=ref["reference_time"],
            reference_score=int(ref["reference_score"]),
        )
    except Exception as exc:
        logger.warning(f"Could not build ITRA predictor from race config: {exc}")
        return None


def _append_pacing_sheet(
    output_path: Path,
    pacing_df,
    sheet_name: str,
    mode: str,
    athlete_name: str,
    itra_score: int | None,
) -> None:
    """Add (or replace) a sheet with the pacing plan in an existing xlsx file."""
    wb = load_workbook(output_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    cols = list(pacing_df.columns)

    # Header
    for col_idx, col_name in enumerate(cols, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data
    for row_idx, row in enumerate(pacing_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Column widths
    col_widths = {
        "Point Name": 30,
        "Total Distance (km)": 18,
        "Elevation (m)": 14,
        "Accum. Elevation Gain (m)": 24,
        "Segment Distance (km)": 20,
        "Segment Elevation Gain (m)": 24,
        "Segment Elevation Loss (m)": 24,
        "Segment Gain (%)": 14,
        "Segment Running Time": 20,
        "Avg Pace (mm:ss/km)": 20,
        "Avg Grade-Adjusted Pace (mm:ss/km)": 28,
        "Stop Time": 12,
        "Elapsed Time": 14,
    }
    for col_idx, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 16)

    # Summary block below the data table
    attrs = pacing_df.attrs
    summary_row = len(pacing_df) + 3
    summary = [
        ("Planning mode", mode),
        ("Athlete", athlete_name),
        ("Total running time", seconds_to_hms(attrs.get("total_running_time_s", 0))),
        ("Overall avg pace", attrs.get("overall_avg_pace_mmss", "-")),
        (
            "Overall avg grade-adjusted pace",
            attrs.get("overall_avg_grade_adjusted_pace_mmss", "-"),
        ),
        ("Total stop time", seconds_to_hms(attrs.get("total_stop_time_s", 0))),
        ("Total finish time", seconds_to_hms(attrs.get("total_time_s", 0))),
    ]
    if attrs.get("fatigue_total_decay_pct", 0) > 0:
        summary.append(
            (
                "Fatigue model",
                f"Linear decay {attrs['fatigue_total_decay_pct']:.1f}%",
            )
        )
    if itra_score is not None:
        summary.append(("Predicted ITRA score", itra_score))

    for offset, (label, value) in enumerate(summary):
        ws.cell(row=summary_row + offset, column=1, value=label)
        ws.cell(row=summary_row + offset, column=2, value=value)

    wb.save(output_path)


def _format_allocations(allocations: list[dict], custom_as_carbs: bool = False) -> str:
    if not allocations:
        return "-"

    rendered = []
    for allocation in allocations:
        units = float(allocation.get("units", 0.0))
        if abs(units) < 1e-9:
            continue
        reference_size = str(allocation.get("reference_size", "")).strip()
        if custom_as_carbs and reference_size.lower() == "custom":
            carbs_g = float(allocation.get("actual_carbs_g", 0.0))
            rendered.append(f"{allocation['food']}: {format_decimal_quantity(carbs_g, 1)} gr CH")
            continue
        rendered.append(
            f"{allocation['food']}: {format_decimal_quantity(units)} x {reference_size}"
        )

    return "; ".join(rendered) if rendered else "-"


def _format_segment_caffeine_events(events: list[dict]) -> str:
    if not events:
        return "-"

    parts = []
    for event in events:
        dose_mg = float(event.get("dose_mg", 0.0))
        if abs(dose_mg) < 1e-9:
            continue
        time_h = float(event.get("time_h", 0.0))
        time_hms = seconds_to_hms(int(round(time_h * 3600.0)))
        parts.append(f"{format_decimal_quantity(dose_mg)} [{time_hms}]")

    return "; ".join(parts) if parts else "-"


def _allocation_category(food_name: str) -> str:
    value = food_name.strip().lower()
    if "water" in value:
        return "water"
    if any(token in value for token in ("pocari", "isotonic", "sports drink")):
        return "sports_drink"
    if any(token in value for token in ("gel", "jelly", "medallist")):
        return "gels"
    return "others"


def _summarize_row_intake_categories(row: dict) -> dict[str, object]:
    allocations = list(row.get("segment_allocations", [])) + list(row.get("aid_allocations", []))
    sports_drink_ml = 0.0
    gel_allocations: list[dict] = []
    other_allocations: list[dict] = []

    for allocation in allocations:
        food_name = str(allocation.get("food", "")).strip()
        category = _allocation_category(food_name)
        if category == "sports_drink":
            volume_ml = extract_volume_ml(str(allocation.get("reference_size", "")))
            sports_drink_ml += float(allocation.get("units", 0.0)) * volume_ml
        elif category == "gels":
            gel_allocations.append(allocation)
        elif category == "others":
            other_allocations.append(allocation)

    # Keep water column as carried/plain water only.
    plain_water_ml = float(row.get("row_supplemental_fluids_ml", 0.0))

    others_text = _format_allocations(other_allocations, custom_as_carbs=True)

    return {
        "water_ml": plain_water_ml,
        "sports_drink_ml": sports_drink_ml,
        "gels_text": _format_allocations(gel_allocations),
        "others_text": others_text,
    }


def _collect_dropbag_points(nutrition_cfg: dict, aid_stations: list[dict]) -> list[str]:
    configured = nutrition_cfg.get("dropbag_points")
    if isinstance(configured, list) and configured:
        return [str(name).strip() for name in configured if str(name).strip()]

    detected: list[str] = []
    for aid in aid_stations or []:
        notes = str(aid.get("notes", ""))
        name = str(aid.get("name", "")).strip()
        if name and "dropbag" in notes.lower():
            detected.append(name)
    return detected


def _build_dropbag_plan(nutrition_plan: dict, dropbag_points: list[str]) -> list[dict]:
    rows = nutrition_plan.get("rows", [])
    if not rows or not dropbag_points:
        return []

    exact_to_index: dict[str, int] = {}
    canonical_to_index: dict[str, int] = {}
    for index, row in enumerate(rows):
        point_name = str(row.get("point_name", "")).strip()
        exact_to_index[point_name] = index
        canonical_to_index[canonical_point_name(point_name)] = index

    bag_indices: list[tuple[str, int]] = []
    for name in dropbag_points:
        raw_name = str(name).strip()
        idx = exact_to_index.get(raw_name)
        if idx is None:
            idx = canonical_to_index.get(raw_name)
        if idx is None:
            idx = canonical_to_index.get(canonical_point_name(raw_name))
        if idx is not None:
            bag_indices.append((name, idx))

    if not bag_indices:
        return []

    bag_indices.sort(key=lambda item: item[1])
    plans: list[dict] = []

    _, first_bag_idx = bag_indices[0]
    if first_bag_idx > 0:
        bag_indices.insert(0, ("START", -1))

    for i, (bag_name, start_idx) in enumerate(bag_indices):
        end_idx = bag_indices[i + 1][1] if i + 1 < len(bag_indices) else len(rows) - 1
        if end_idx <= start_idx:
            continue

        by_food: dict[str, dict[str, float | str]] = {}
        caffeine_events: list[dict] = []
        for row in rows[start_idx + 1 : end_idx + 1]:
            for allocation in row.get("segment_allocations", []):
                units = float(allocation.get("units", 0.0))
                if abs(units) < 1e-9:
                    continue
                food = str(allocation.get("food", "")).strip()
                if not food:
                    continue
                stats = by_food.setdefault(
                    food,
                    {
                        "units": 0.0,
                        "reference_size": str(allocation.get("reference_size", "1 unit")),
                    },
                )
                stats["units"] = float(stats["units"]) + units

            caffeine_events.extend(row.get("segment_caffeine_events", []))

        formatted_allocations = [
            {
                "food": food,
                "units": values["units"],
                "reference_size": values["reference_size"],
            }
            for food, values in sorted(by_food.items())
        ]

        plans.append(
            {
                "dropbag_point": bag_name,
                "covers_until": rows[end_idx].get("point_name", ""),
                "food_allocations": formatted_allocations,
                "caffeine_events": sorted(
                    caffeine_events,
                    key=lambda event: float(event.get("time_h", 0.0)),
                ),
            }
        )

    return plans


def _append_dropbag_sheet(output_path: Path, dropbag_plan: list[dict], sheet_name: str) -> None:
    wb = load_workbook(output_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    headers = [
        "Dropbag Point",
        "Covers Segments Until",
        "Stash Food Plan (qty)",
        "Stash Caffeine Plan (mg [time])",
    ]
    for col_idx, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=name)

    for row_idx, row in enumerate(dropbag_plan, start=2):
        ws.cell(row=row_idx, column=1, value=row.get("dropbag_point", ""))
        ws.cell(row=row_idx, column=2, value=row.get("covers_until", ""))
        ws.cell(row=row_idx, column=3, value=_format_allocations(row.get("food_allocations", [])))
        ws.cell(
            row=row_idx,
            column=4,
            value=_format_segment_caffeine_events(row.get("caffeine_events", [])),
        )

    widths = {1: 26, 2: 30, 3: 70, 4: 34}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(output_path)


def _append_nutrition_sheet(output_path: Path, nutrition_plan: dict, sheet_name: str) -> None:
    """Add (or replace) a sheet with nutrition quantities in the race workbook."""
    wb = load_workbook(output_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    headers = [
        "Point Name",
        "Segment Time",
        "Elapsed Time",
        "Segment Carb Target (g)",
        "Moving Plan Carbs (g)",
        "Aid Plan Carbs (g)",
        "Total Segment Carbs per Hour (g/h)",
        "Estimated Sweat Loss (ml)",
        "Total Planned Drink (ml)",
        "Cum Sweat Imbalance (ml)",
        "Cum Sweat Imbalance (%BW)",
        "Segment Caffeine Intake (mg [time])",
        "Caffeine Concentration (mg/kg)",
        "Water (ml)",
        "Sports Drink (ml)",
        "Gels (qty)",
        "Others (qty)",
    ]
    for col_idx, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=name)

    for row_idx, row in enumerate(nutrition_plan["rows"], start=2):
        category_summary = _summarize_row_intake_categories(row)
        row_values = [
            row["point_name"],
            row["segment_time_hms"],
            row["elapsed_hms"],
            round(row["row_target_carbs_g"], 1),
            round(row["segment_target_carbs_g"], 1),
            round(row["aid_target_carbs_g"], 1),
            round(row["row_carbs_per_h"], 1),
            round(row.get("row_sweat_loss_ml", 0.0), 1),
            round(row.get("row_total_fluids_ml", 0.0), 1),
            round(row.get("cumulative_hydration_balance_ml", 0.0), 1),
            round(row.get("cumulative_hydration_balance_pct_bw", 0.0), 2),
            _format_segment_caffeine_events(row.get("segment_caffeine_events", [])),
            round(row.get("caffeine_concentration_mg_per_kg", 0.0), 2),
            round(float(category_summary["water_ml"]), 1),
            round(float(category_summary["sports_drink_ml"]), 1),
            str(category_summary["gels_text"]),
            str(category_summary["others_text"]),
        ]
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    widths = {
        1: 30,
        2: 14,
        3: 14,
        4: 22,
        5: 20,
        6: 18,
        7: 32,
        8: 20,
        9: 20,
        10: 24,
        11: 24,
        12: 34,
        13: 28,
        14: 14,
        15: 18,
        16: 50,
        17: 58,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    totals = nutrition_plan["totals"]
    summary_start = len(nutrition_plan["rows"]) + 3
    summary_rows = [
        ("Target carbs per hour (g/h)", round(totals["target_carbs_g_per_h"], 1)),
        ("Planned duration (h)", round(totals["total_time_h"], 2)),
        ("Moving target carbs (g)", round(totals["moving_target_carbs_g"], 1)),
        ("Planned total carbs (segment + aid) (g)", round(totals["planned_total_carbs_g"], 1)),
        (
            "Sweat loss estimate total (ml)",
            round(totals.get("hydration", {}).get("estimated_total_sweat_loss_ml", 0.0), 1),
        ),
        (
            "Planned drink total (ml)",
            round(totals.get("hydration", {}).get("planned_total_fluids_ml", 0.0), 1),
        ),
        (
            "Final sweat imbalance (ml)",
            round(totals.get("hydration", {}).get("final_sweat_imbalance_ml", 0.0), 1),
        ),
        (
            "Final sweat imbalance (%BW)",
            round(totals.get("hydration", {}).get("final_sweat_imbalance_pct_bw", 0.0), 2),
        ),
        (
            "Caffeine total dose (mg)",
            round(totals.get("caffeine", {}).get("total_dose_mg", 0.0), 1),
        ),
        (
            "Caffeine peak concentration (mg/kg)",
            round(totals.get("caffeine", {}).get("peak_concentration_mg_per_kg", 0.0), 2),
        ),
    ]
    for offset, (label, value) in enumerate(summary_rows):
        ws.cell(row=summary_start + offset, column=1, value=label)
        ws.cell(row=summary_start + offset, column=2, value=value)

    food_rows_start = summary_start + len(summary_rows) + 2
    ws.cell(row=food_rows_start, column=1, value="Food Totals")
    ws.cell(row=food_rows_start, column=2, value="Units")
    ws.cell(row=food_rows_start, column=3, value="Carbs (g)")

    for offset, food_name in enumerate(sorted(totals["by_food"].keys()), start=1):
        stats = totals["by_food"][food_name]
        ws.cell(row=food_rows_start + offset, column=1, value=food_name)
        ws.cell(row=food_rows_start + offset, column=2, value=round(stats["units"], 2))
        ws.cell(row=food_rows_start + offset, column=3, value=round(stats["carbs_g"], 1))

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Trail race planner — segment analysis and pacing plan"
    )
    parser.add_argument("race_config", type=Path, help="Path to race YAML config")
    parser.add_argument(
        "--athlete",
        default="yet_another_sato",
        help="Athlete name (default: yet_another_sato)",
    )
    parser.add_argument(
        "--mode",
        choices=["athlete_pb", "target_time", "target_itra", "grade_adjusted_pace"],
        default="athlete_pb",
        help="Planning mode (default: athlete_pb)",
    )
    parser.add_argument(
        "--target-time",
        metavar="HH:MM:SS",
        help="Desired total finish time — required for --mode target_time",
    )
    parser.add_argument(
        "--target-itra-score",
        type=int,
        metavar="N",
        help="Target ITRA score — required for --mode target_itra",
    )
    parser.add_argument(
        "--target-grade-adjusted-pace",
        metavar="MM:SS",
        help=(
            "Target grade-adjusted running pace in MM:SS or MM:SS/km "
            "— required for --mode grade_adjusted_pace"
        ),
    )
    parser.add_argument(
        "--fatigue-mode",
        choices=["none", "athlete", "race"],
        default="none",
        help="Fatigue model source (default: none — no fatigue)",
    )
    parser.add_argument(
        "--fatigue-total-decay-pct",
        type=float,
        metavar="PCT",
        help="Override fatigue model with linear decay PCT (0–100); takes precedence over config",
    )
    parser.add_argument(
        "--nutrition",
        choices=["yes", "no"],
        default="no",
        help="Include nutrition column in main HTML report (default: no)",
    )
    args = parser.parse_args()

    if args.mode == "target_time" and not args.target_time:
        parser.error("--target-time HH:MM:SS is required when --mode target_time")
    if args.mode == "target_itra" and not args.target_itra_score:
        parser.error("--target-itra-score N is required when --mode target_itra")
    if args.mode == "grade_adjusted_pace" and not args.target_grade_adjusted_pace:
        parser.error(
            "--target-grade-adjusted-pace MM:SS is required when --mode grade_adjusted_pace"
        )

    # Validate fatigue arguments
    if args.fatigue_total_decay_pct is not None:
        if not 0 <= args.fatigue_total_decay_pct <= 100:
            parser.error("--fatigue-total-decay-pct must be between 0 and 100")

    # ------------------------------------------------------------------
    # Load configs
    # ------------------------------------------------------------------
    race_config_path = args.race_config
    if not race_config_path.exists():
        logger.error(f"Race configuration file not found: {race_config_path}")
        sys.exit(1)

    project_root = Path.cwd()
    athlete_config_path = project_root / "config" / "athletes" / f"{args.athlete}.yaml"
    if not athlete_config_path.exists():
        logger.error(f"Athlete configuration file not found: {athlete_config_path}")
        sys.exit(1)

    logger.info(f"Race config:    {race_config_path}")
    logger.info(f"Athlete config: {athlete_config_path}")
    logger.info(f"Planning mode:  {args.mode}")

    with open(race_config_path, "r", encoding="utf-8") as f:
        race_config = yaml.safe_load(f)
    with open(athlete_config_path, "r", encoding="utf-8") as f:
        athlete_config = yaml.safe_load(f)

    race_info = race_config.get("race", {})
    gpx_file = race_info.get("gpx_file")
    output_file = race_info.get("output_file")
    resample_m = race_info.get("resample_m", 5)

    if not gpx_file:
        logger.error("'gpx_file' not specified in race config")
        sys.exit(1)
    if not output_file:
        logger.error("'output_file' not specified in race config")
        sys.exit(1)

    gpx_path = project_root / gpx_file
    output_path = project_root / output_file

    athlete_info = athlete_config.get("athlete", {})
    athlete_display_name = athlete_info.get("name", args.athlete)
    logger.info(f"Athlete: {athlete_display_name}")

    # ------------------------------------------------------------------
    # Resolve fatigue configuration (CLI > race > athlete > 0)
    # ------------------------------------------------------------------
    fatigue_total_decay_pct = _resolve_fatigue_total_decay_pct(
        fatigue_mode=args.fatigue_mode,
        fatigue_total_decay_pct_cli=args.fatigue_total_decay_pct,
        race_config=race_config,
        athlete_config=athlete_config,
    )
    if fatigue_total_decay_pct > 0:
        logger.info(f"Fatigue model: linear decay {fatigue_total_decay_pct:.1f}%")

    # ------------------------------------------------------------------
    # 1. Segment analysis — always runs; creates / updates the xlsx file
    # ------------------------------------------------------------------
    try:
        analyzer = analyze_course(
            gpx_path=gpx_path,
            race_config_path=race_config_path,
            output_path=output_path,
            resample_m=resample_m,
            athlete_config=athlete_config,
        )
        logger.success(f"Segment analysis written to: {output_path}")
    except Exception as exc:
        logger.error(f"Segment analysis failed: {exc}")
        import traceback

        traceback.print_exc()
        sys.exit(1)

    course = analyzer.course
    aid_stations = analyzer.aid_stations

    # ------------------------------------------------------------------
    # 2. ITRA predictor — built here so it's available for both
    #    target_itra mode and score output at the end
    # ------------------------------------------------------------------
    itra_predictor = _build_itra_predictor(race_config)

    # ------------------------------------------------------------------
    # 3. Resolve target running time and build PaceCalculator
    # ------------------------------------------------------------------
    override_running_time_s: float | None = None

    if args.mode == "athlete_pb":
        calc = PaceCalculator.from_athlete_config(
            athlete_config, fatigue_total_decay_pct=fatigue_total_decay_pct
        )
        ref = athlete_info.get("reference_performance", {})
        logger.info(f"Reference performance: {ref.get('distance_km')} km in {ref.get('time')}")

    elif args.mode == "target_time":
        target_total_s = float(hms_to_seconds(args.target_time))
        total_stop_s = _total_stop_time_s(aid_stations)
        override_running_time_s = target_total_s - total_stop_s
        if override_running_time_s <= 0:
            logger.error(
                f"Target time {args.target_time} is shorter than total stop time "
                f"({seconds_to_hms(total_stop_s)})"
            )
            sys.exit(1)
        calc = PaceCalculator.from_athlete_config(
            athlete_config, fatigue_total_decay_pct=fatigue_total_decay_pct
        )
        logger.info(
            f"Target finish time: {args.target_time}  "
            f"(running: {seconds_to_hms(override_running_time_s)}, "
            f"stops: {seconds_to_hms(total_stop_s)})"
        )

    elif args.mode == "target_itra":
        if itra_predictor is None:
            logger.error(
                "Cannot use --mode target_itra: " "no 'itra_reference_points' in race config."
            )
            sys.exit(1)
        predicted_total_time_h = itra_predictor.predict_time(args.target_itra_score)
        target_total_time_s = predicted_total_time_h * 3600
        total_stop_s = _total_stop_time_s(aid_stations)
        override_running_time_s = target_total_time_s - total_stop_s
        if override_running_time_s <= 0:
            logger.error(
                f"ITRA-predicted time {hours_to_hms(predicted_total_time_h)} is shorter "
                f"than total stop time ({seconds_to_hms(total_stop_s)})"
            )
            sys.exit(1)
        calc = PaceCalculator.from_athlete_config(
            athlete_config, fatigue_total_decay_pct=fatigue_total_decay_pct
        )
        logger.info(
            f"ITRA score {args.target_itra_score} → "
            f"predicted finish time: {hours_to_hms(predicted_total_time_h)}  "
            f"(running: {seconds_to_hms(override_running_time_s)}, "
            f"stops: {seconds_to_hms(total_stop_s)})"
        )

    elif args.mode == "grade_adjusted_pace":
        calc = PaceCalculator.from_athlete_config(
            athlete_config, fatigue_total_decay_pct=fatigue_total_decay_pct
        )
        try:
            target_grade_adjusted_pace_s_per_km = pace_to_seconds_per_km(
                args.target_grade_adjusted_pace
            )
        except ValueError as exc:
            logger.error(f"Invalid --target-grade-adjusted-pace value: {exc}")
            sys.exit(1)

        if target_grade_adjusted_pace_s_per_km <= 0:
            logger.error("--target-grade-adjusted-pace must be positive")
            sys.exit(1)

        planned_finish_distance_km = (
            float(aid_stations[-1].get("distance_km", course.total_distance_km))
            if aid_stations
            else None
        )
        total_grade_weighted_km = calc.grade_weighted_distance_km(
            course,
            end_distance_km=planned_finish_distance_km,
        )
        override_running_time_s = total_grade_weighted_km * target_grade_adjusted_pace_s_per_km
        total_stop_s = _total_stop_time_s(aid_stations)
        logger.info(
            f"Target grade-adjusted pace: {args.target_grade_adjusted_pace}  "
            f"(weighted distance: {total_grade_weighted_km:.2f} km, "
            f"running: {seconds_to_hms(override_running_time_s)}, "
            f"stops: {seconds_to_hms(total_stop_s)})"
        )

    # ------------------------------------------------------------------
    # 4. Pacing plan
    # ------------------------------------------------------------------
    try:
        pacing_df = calc.calculate_pacing(
            course=course,
            aid_stations=aid_stations,
            use_fed=(args.mode == "athlete_pb"),
            override_total_running_time_s=override_running_time_s,
        )
    except Exception as exc:
        logger.error(f"Pacing calculation failed: {exc}")
        import traceback

        traceback.print_exc()
        sys.exit(1)

    total_time_s = pacing_df.attrs["total_time_s"]

    # ------------------------------------------------------------------
    # 5. ITRA score for the computed finish time
    # ------------------------------------------------------------------
    itra_score_result: int | None = None
    if itra_predictor is not None:
        itra_score_result = itra_predictor.predict_score(total_time_s / 3600)
        logger.info(
            f"Expected ITRA score for finish time "
            f"{seconds_to_hms(total_time_s)}: {itra_score_result}"
        )
    else:
        logger.warning("ITRA score output skipped (no reference data in race config).")

    # ------------------------------------------------------------------
    # 6. Append pacing sheet to xlsx
    # ------------------------------------------------------------------
    sheet_name = f"Race Plan ({args.mode})"
    _append_pacing_sheet(
        output_path=output_path,
        pacing_df=pacing_df,
        sheet_name=sheet_name,
        mode=args.mode,
        athlete_name=athlete_display_name,
        itra_score=itra_score_result,
    )
    logger.success(f"Pacing plan written to sheet '{sheet_name}' in {output_path}")

    # ------------------------------------------------------------------
    # 7. Nutrition plan sheet (optional)
    # ------------------------------------------------------------------
    carb_plan: dict | None = None
    nutrition_cfg = race_config.get("nutrition")
    if nutrition_cfg:
        try:
            target_carbs_g_per_h = float(nutrition_cfg.get("target_carbs_g_per_h"))
            catalog_file = nutrition_cfg.get("food_catalog_file", "config/nutrition/foods.yaml")
            catalog_path = project_root / catalog_file
            if not catalog_path.exists():
                raise FileNotFoundError(f"Food catalog file not found: {catalog_path}")

            with open(catalog_path, "r", encoding="utf-8") as f:
                catalog_cfg = yaml.safe_load(f)
            food_catalog = load_food_catalog(catalog_cfg)

            caffeine_ingestion_plan = [
                (float(entry.get("time_h", 0.0)), float(entry.get("dose_mg", 0.0)))
                for entry in nutrition_cfg.get("caffeine_plan", {}).get("ingestion_plan", [])
            ]
            athlete_weight_kg = float(
                athlete_config.get("athlete", {}).get("weight_kg", 0.0) or 0.0
            )
            if caffeine_ingestion_plan and athlete_weight_kg <= 0:
                raise ValueError(
                    "Caffeine plan is configured but athlete.weight_kg is missing or invalid"
                )

            athlete_hydration_cfg = athlete_config.get("athlete", {}).get("hydration", {}) or {}
            sweat_rate_ml_per_h = athlete_hydration_cfg.get("sweat_rate_ml_per_h")
            if sweat_rate_ml_per_h is not None:
                sweat_rate_ml_per_h = float(sweat_rate_ml_per_h)
                if sweat_rate_ml_per_h < 0:
                    raise ValueError("athlete.hydration.sweat_rate_ml_per_h must be >= 0")
                if sweat_rate_ml_per_h > 0 and athlete_weight_kg <= 0:
                    raise ValueError(
                        "Hydration plan is configured but athlete.weight_kg is missing or invalid"
                    )

            carb_plan = build_race_nutrition_plan(
                pacing_rows=pacing_df.to_dict(orient="records"),
                target_carbs_g_per_h=target_carbs_g_per_h,
                food_catalog=food_catalog,
                segment_foods_cfg=nutrition_cfg.get("segment_foods", {}),
                aid_station_intake_cfg=nutrition_cfg.get("aid_station_intake", {}),
                caffeine_ingestion_plan=caffeine_ingestion_plan,
                caffeine_weight_kg=athlete_weight_kg,
                sweat_rate_ml_per_h=sweat_rate_ml_per_h,
                hydration_weight_kg=athlete_weight_kg,
            )
            _append_nutrition_sheet(
                output_path=output_path,
                nutrition_plan=carb_plan,
                sheet_name="Nutrition Plan",
            )
            logger.success("Nutrition plan written to sheet 'Nutrition Plan'")

            dropbag_points = _collect_dropbag_points(nutrition_cfg, aid_stations)
            dropbag_plan = _build_dropbag_plan(carb_plan, dropbag_points)
            if dropbag_plan:
                _append_dropbag_sheet(
                    output_path=output_path,
                    dropbag_plan=dropbag_plan,
                    sheet_name="Dropbag Plan",
                )
                logger.success("Dropbag plan written to sheet 'Dropbag Plan'")
        except Exception as exc:
            logger.error(f"Nutrition plan generation failed: {exc}")

    # ------------------------------------------------------------------
    # 8. Generate smartphone-friendly race plan HTML report
    # ------------------------------------------------------------------
    try:
        race_name = race_info.get("name", "Race Plan")
        race_start_time = race_info.get("start_time")
        report_stem = output_path.stem.removesuffix("_segment_analysis")
        html_output_path = output_path.parent / f"{report_stem}_race_plan.html"
        nutrition_plan_for_html = carb_plan if args.nutrition == "yes" else None
        generate_race_plan_table_report(
            course=course,
            aid_stations=aid_stations,
            pacing_df=pacing_df,
            output_path=html_output_path,
            race_name=race_name,
            mode=args.mode,
            nutrition_plan=nutrition_plan_for_html,
            race_start_time=race_start_time,
            title=f"{race_name} – Race Plan",
        )
        logger.success(f"Race plan HTML report written to: {html_output_path}")
    except Exception as exc:
        logger.error(f"Race plan HTML report generation failed: {exc}")
        import traceback

        traceback.print_exc()

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------
    sep = "=" * 52
    logger.info(sep)
    logger.info("RACE PLAN SUMMARY")
    logger.info(f"  Mode:          {args.mode}")
    logger.info(f"  Athlete:       {athlete_display_name}")
    approx_running_s = pacing_df.attrs.get("riegel_running_time_approx_s")
    if approx_running_s is not None:
        logger.info(f"  Riegel approx running time: {seconds_to_hms(float(approx_running_s))}")
        logger.info(
            f"  Grade-adjusted running time: "
            f"{seconds_to_hms(pacing_df.attrs['total_running_time_s'])}"
        )
    else:
        logger.info(
            f"  Running time:  " f"{seconds_to_hms(pacing_df.attrs['total_running_time_s'])}"
        )
    logger.info(f"  Stop time:     {seconds_to_hms(pacing_df.attrs['total_stop_time_s'])}")
    logger.info(f"  Avg pace:      {pacing_df.attrs.get('overall_avg_pace_mmss', '-')}/km")
    logger.info(
        "  Avg grade-adjusted pace: "
        f"{pacing_df.attrs.get('overall_avg_grade_adjusted_pace_mmss', '-')}/km"
    )
    if pacing_df.attrs.get("fatigue_total_decay_pct", 0) > 0:
        logger.info(
            f"  Fatigue model:  Linear decay {pacing_df.attrs['fatigue_total_decay_pct']:.1f}%"
        )
    logger.info(f"  Finish time:   {seconds_to_hms(total_time_s)}")
    if itra_score_result is not None:
        logger.info(f"  ITRA score:    {itra_score_result}")
    logger.info(sep)


if __name__ == "__main__":
    main()
